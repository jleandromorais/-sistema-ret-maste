"""
════════════════════════════════════════════════════════════════════════════════
    SISTEMA RET MASTER - PROCESSAMENTO AUTOMATIZADO DE PDFs
════════════════════════════════════════════════════════════════════════════════

📚 ARQUIVO DE ESTUDO - VERSÃO COMENTADA

Este arquivo contém o código completo do Sistema RET Master com comentários
detalhados para facilitar o entendimento de cada parte do sistema.

🎯 OBJETIVO:
    Processar PDFs de RET (EAT, Penalidades, TOP), extrair dados estruturados
    e gerar relatórios em Excel formatado + banco de dados SQLite.

🔧 TECNOLOGIAS USADAS:
    - CustomTkinter: Interface gráfica moderna
    - PDFPlumber: Extração de texto de PDFs
    - Pandas: Manipulação de dados
    - OpenPyXL: Formatação de Excel
    - SQLite3: Banco de dados local
    - Regex: Extração de padrões de texto

📖 ESTRUTURA DO CÓDIGO:
    1. Importações e Configurações Globais
    2. Classe Principal (SistemaRET)
       - Inicialização
       - Setup da Interface (UI)
       - Funções de Processamento
       - Funções de Exportação
    3. Execução Principal

════════════════════════════════════════════════════════════════════════════════
"""

# ═══════════════════════════════════════════════════════════════════════════
# 📦 SEÇÃO 1: IMPORTAÇÕES
# ═══════════════════════════════════════════════════════════════════════════

# --- Bibliotecas do Sistema Operacional ---
import os                    # Manipulação de arquivos e caminhos
import sqlite3              # Banco de dados SQLite

# --- Análise e Manipulação de Dados ---
import pandas as pd         # Análise de dados em DataFrame (como Excel no código)

# --- Interface Gráfica ---
import customtkinter as ctk              # Framework moderno para GUI
from tkinter import filedialog, messagebox  # Diálogos nativos do SO

# --- Processamento de PDFs ---
import pdfplumber           # Biblioteca para extrair texto de PDFs

# --- Utilidades ---
import re                   # Expressões regulares (regex) para padrões de texto
from datetime import datetime           # Data e hora atual

# --- Formatação de Excel ---
from openpyxl import Workbook                                    # Criar arquivo Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Estilos
from openpyxl.utils.dataframe import dataframe_to_rows          # Converter DataFrame para Excel


# ═══════════════════════════════════════════════════════════════════════════
# ⚙️ SEÇÃO 2: CONFIGURAÇÕES GLOBAIS
# ═══════════════════════════════════════════════════════════════════════════

# --- Configuração do Tema Visual ---
ctk.set_appearance_mode("Dark")         # Define modo escuro (Dark/Light)
ctk.set_default_color_theme("blue")     # Define tema de cores (blue/green/dark-blue)


# ═══════════════════════════════════════════════════════════════════════════
# 🏗️ SEÇÃO 3: CLASSE PRINCIPAL DO SISTEMA
# ═══════════════════════════════════════════════════════════════════════════

class SistemaRET(ctk.CTk):
    """
    📌 CLASSE PRINCIPAL DO SISTEMA RET MASTER
    
    Esta classe herda de CTk (CustomTkinter) e representa a janela principal
    do aplicativo. Ela contém toda a lógica de interface e processamento.
    
    🔧 ATRIBUTOS:
        - pasta_selecionada: Caminho da pasta a ser processada
        - dados_processados: Lista com todos os dados extraídos dos PDFs
        - resultados: Resumo estatístico do processamento
        - tipos_encargo: Dicionário com checkboxes dos tipos
        
    📝 MÉTODOS PRINCIPAIS:
        - __init__(): Construtor da classe
        - _setup_ui(): Cria toda a interface visual
        - selecionar_pasta(): Abre diálogo de seleção de pasta
        - processar(): Processa todos os PDFs
        - exportar_excel(): Gera arquivo Excel formatado
        - salvar_db(): Salva dados no SQLite
    """
    
    def __init__(self):
        """
        🎬 CONSTRUTOR DA CLASSE
        
        Executa quando criamos um objeto SistemaRET().
        Inicializa a janela principal e todas as variáveis.
        """
        # Chama o construtor da classe pai (CTk)
        super().__init__()
        
        # --- Configurações da Janela Principal ---
        self.title("Sistema RET - Processamento de PDFs")  # Título da janela
        self.geometry("1400x900")                          # Tamanho (largura x altura)
        
        # --- Inicialização das Variáveis de Dados ---
        self.pasta_selecionada = None        # Ainda não selecionou pasta
        self.dados_processados = []          # Lista vazia para armazenar dados
        self.resultados = None               # Será preenchido após processar
        
        # --- Criar Interface ---
        self._setup_ui()  # Chama método que cria todos os componentes visuais
    
    
    def _setup_ui(self):
        """
        🎨 CONFIGURAÇÃO DA INTERFACE GRÁFICA
        
        Este método cria TODOS os componentes visuais da janela:
        - Header (cabeçalho)
        - Painel esquerdo (controles)
        - Painel direito (resultados)
        - Rodapé (ações e total)
        
        💡 DICA: O CustomTkinter usa o sistema de "pack" para organizar elementos.
        """
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📋 HEADER (CABEÇALHO SUPERIOR)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        # Criar frame do header com cor escura
        header = ctk.CTkFrame(
            self,                           # Pai = janela principal
            height=80,                      # Altura fixa de 80 pixels
            corner_radius=0,                # Sem bordas arredondadas
            fg_color="#1a1a2e"             # Cor de fundo azul escuro
        )
        header.pack(fill="x")               # Preenche horizontalmente
        header.pack_propagate(False)        # Mantém altura fixa
        
        # Label do título principal
        ctk.CTkLabel(
            header,
            text="Sistema RET Master",
            font=("Roboto", 32, "bold"),    # Fonte grande e negrito
            text_color="#00d9ff"            # Cor ciano brilhante
        ).pack(side="left", padx=30, pady=20)
        
        # Label do subtítulo
        ctk.CTkLabel(
            header,
            text="Processamento Automatizado de Encargos",
            font=("Roboto", 14),
            text_color="#a0a0a0"            # Cor cinza claro
        ).pack(side="left", padx=10)
        
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📦 CONTAINER PRINCIPAL (CORPO DA JANELA)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        main = ctk.CTkFrame(self, fg_color="transparent")  # Frame transparente
        main.pack(fill="both", expand=True, padx=20, pady=20)
        
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # ◀️ PAINEL ESQUERDO (CONTROLES E SELEÇÃO)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        left = ctk.CTkFrame(main, width=400, corner_radius=15)
        left.pack(side="left", fill="both", padx=(0, 10), pady=0)
        left.pack_propagate(False)  # Mantém largura fixa de 400px
        
        # --- Título da Seção ---
        ctk.CTkLabel(
            left,
            text="Seleção de Pasta",
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        # --- Label que mostra a pasta selecionada ---
        self.lbl_pasta = ctk.CTkLabel(
            left,
            text="Nenhuma pasta selecionada",
            font=("Roboto", 12),
            wraplength=350,                 # Quebra texto em 350px
            text_color="#808080"            # Cinza
        )
        self.lbl_pasta.pack(pady=10, padx=20)
        
        # --- Botão para Selecionar Pasta ---
        ctk.CTkButton(
            left,
            text="Selecionar Pasta",
            command=self.selecionar_pasta,  # ← Chama função quando clica
            height=40,
            font=("Roboto", 14, "bold"),
            fg_color="#2196F3",             # Azul
            hover_color="#1976D2"           # Azul mais escuro no hover
        ).pack(pady=10, padx=20, fill="x")
        
        # --- Linha Separadora ---
        ctk.CTkFrame(
            left,
            height=2,
            fg_color="#404040"              # Cinza escuro
        ).pack(fill="x", pady=20, padx=20)
        
        # --- Título dos Tipos de Encargo ---
        ctk.CTkLabel(
            left,
            text="Tipos de Encargo",
            font=("Roboto", 18, "bold")
        ).pack(pady=(10, 5), padx=20, anchor="w")
        
        # --- Checkboxes para Tipos de Encargo ---
        # 💡 Armazenamos em um dicionário para facilitar acesso depois
        self.tipos_encargo = {
            "EAT": ctk.CTkCheckBox(
                left,
                text="EAT (Encargos de Acesso e Transporte)"
            ),
            "Penalidades": ctk.CTkCheckBox(
                left,
                text="Penalidades"
            ),
            "TOP": ctk.CTkCheckBox(
                left,
                text="TOP (Takeoff Point)"
            )
        }
        
        # Marca todos como selecionados por padrão e exibe na tela
        for checkbox in self.tipos_encargo.values():
            checkbox.select()               # ← Marca checkbox
            checkbox.pack(pady=5, padx=30, anchor="w")
        
        # --- Botão PROCESSAR (Principal) ---
        ctk.CTkButton(
            left,
            text="⚡ PROCESSAR PDFs",
            command=self.processar,          # ← Chama função processar()
            height=50,
            font=("Roboto", 16, "bold"),
            fg_color="#4CAF50",              # Verde
            hover_color="#45a049"            # Verde escuro
        ).pack(pady=30, padx=20, fill="x")
        
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # ▶️ PAINEL DIREITO (RESULTADOS E VISUALIZAÇÃO)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        right = ctk.CTkFrame(main, corner_radius=15)
        right.pack(side="right", fill="both", expand=True)
        
        # --- Título da Seção ---
        ctk.CTkLabel(
            right,
            text="Resultados do Processamento",
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        # --- Sistema de Abas (TabView) ---
        # 💡 Permite alternar entre Resumo, Dados e Logs
        self.tabview = ctk.CTkTabview(right)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Adicionar 3 abas
        self.tabview.add("Resumo")
        self.tabview.add("Dados Detalhados")
        self.tabview.add("Logs")
        
        # --- ABA 1: RESUMO ---
        self.frame_resumo = ctk.CTkScrollableFrame(
            self.tabview.tab("Resumo")
        )
        self.frame_resumo.pack(fill="both", expand=True)
        
        # Label para estatísticas
        self.lbl_stats = ctk.CTkLabel(
            self.frame_resumo,
            text="Aguardando processamento...",
            font=("Roboto", 14),
            justify="left"
        )
        self.lbl_stats.pack(pady=20, padx=20, anchor="w")
        
        # --- ABA 2: DADOS DETALHADOS ---
        self.frame_dados = ctk.CTkScrollableFrame(
            self.tabview.tab("Dados Detalhados")
        )
        self.frame_dados.pack(fill="both", expand=True)
        
        # --- ABA 3: LOGS ---
        # 💡 TextBox permite exibir múltiplas linhas de texto
        self.txt_logs = ctk.CTkTextbox(
            self.tabview.tab("Logs"),
            font=("Consolas", 11)           # Fonte monoespaçada
        )
        self.txt_logs.pack(fill="both", expand=True, padx=10, pady=10)
        
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📊 RODAPÉ (TOTAL GERAL E BOTÕES DE AÇÃO)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        footer = ctk.CTkFrame(
            self,
            height=100,
            corner_radius=15,
            fg_color="#1a1a2e"
        )
        footer.pack(fill="x", padx=20, pady=(0, 20))
        footer.pack_propagate(False)
        
        # --- Exibição do Total Geral (Lado Esquerdo) ---
        result_frame = ctk.CTkFrame(footer, fg_color="transparent")
        result_frame.pack(side="left", padx=30, pady=20)
        
        ctk.CTkLabel(
            result_frame,
            text="TOTAL GERAL:",
            font=("Roboto", 14)
        ).pack(anchor="w")
        
        self.lbl_total = ctk.CTkLabel(
            result_frame,
            text="EUR 0,00",
            font=("Roboto", 28, "bold"),
            text_color="#00d9ff"             # Ciano brilhante
        )
        self.lbl_total.pack(anchor="w")
        
        # --- Botões de Ação (Lado Direito) ---
        btn_frame = ctk.CTkFrame(footer, fg_color="transparent")
        btn_frame.pack(side="right", padx=30, pady=20)
        
        # Botão Salvar no Banco
        ctk.CTkButton(
            btn_frame,
            text="💾 Salvar no Banco",
            command=self.salvar_db,
            width=140,
            height=35,
            fg_color="#9C27B0",              # Roxo
            hover_color="#7B1FA2"
        ).pack(side="left", padx=5)
        
        # Botão Exportar Excel
        ctk.CTkButton(
            btn_frame,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            width=140,
            height=35,
            fg_color="#FF9800",              # Laranja
            hover_color="#F57C00"
        ).pack(side="left", padx=5)
    
    
    # ═══════════════════════════════════════════════════════════════════════
    # 📝 FUNÇÕES AUXILIARES DA INTERFACE
    # ═══════════════════════════════════════════════════════════════════════
    
    def log(self, mensagem):
        """
        📋 ADICIONA MENSAGEM AO LOG
        
        Esta função adiciona uma linha no TextBox de logs com timestamp.
        
        Args:
            mensagem (str): Texto a ser adicionado ao log
        """
        timestamp = datetime.now().strftime("%H:%M:%S")  # Hora:Minuto:Segundo
        self.txt_logs.insert("end", f"[{timestamp}] {mensagem}\n")
        self.txt_logs.see("end")    # Scroll automático para final
        self.update()               # Atualiza a interface imediatamente
    
    
    def selecionar_pasta(self):
        """
        📁 ABRE DIÁLOGO PARA SELECIONAR PASTA
        
        Abre janela nativa do SO para o usuário escolher a pasta.
        Atualiza o label com o caminho selecionado.
        """
        # Abre diálogo de seleção
        pasta = filedialog.askdirectory(
            title="Selecione a Pasta Principal (RET)"
        )
        
        if pasta:  # Se usuário não cancelou
            self.pasta_selecionada = pasta
            
            # Atualiza label para mostrar pasta selecionada
            self.lbl_pasta.configure(
                text=f"✓ Pasta: {pasta}",
                text_color="#4CAF50"    # Verde = sucesso
            )
            
            self.log(f"✓ Pasta selecionada: {pasta}")
    
    
    # ═══════════════════════════════════════════════════════════════════════
    # 🔍 FUNÇÕES DE EXTRAÇÃO DE DADOS DOS PDFs
    # ═══════════════════════════════════════════════════════════════════════
    
    def extrair_dados_pdf(self, caminho_pdf):
        """
        📄 EXTRAI DADOS ESTRUTURADOS DE UM PDF
        
        Esta é a função MAIS IMPORTANTE do sistema!
        Ela abre o PDF e usa regex para extrair informações específicas.
        
        Args:
            caminho_pdf (str): Caminho completo do arquivo PDF
            
        Returns:
            dict: Dicionário com todos os dados extraídos
        
        💡 TÉCNICAS USADAS:
            - PDFPlumber: Lê o texto do PDF
            - Regex (re): Encontra padrões específicos no texto
            - Funções auxiliares: Identificam tipo, empresa, etc.
        """
        
        # Inicializa dicionário com dados vazios
        dados = {
            'arquivo': os.path.basename(caminho_pdf),
            'caminho': caminho_pdf,
            'tipo_encargo': self._identificar_tipo(caminho_pdf),
            'empresa': self._extrair_empresa(caminho_pdf),
            'nota_tipo': self._extrair_tipo_nota(caminho_pdf),
            'numero_nd': '',
            'data_vencimento': '',
            'valor_total': 0.0,
            'quantidade': 0.0,
            'valor_unitario': 0.0,
            'valores_encontrados': []
        }
        
        try:
            # Abre o PDF com PDFPlumber
            with pdfplumber.open(caminho_pdf) as pdf:
                texto_completo = ''
                
                # Loop por todas as páginas do PDF
                for pagina in pdf.pages:
                    texto = pagina.extract_text()
                    if texto:
                        texto_completo += texto + '\n'
                
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # 🔍 EXTRAÇÃO 1: NÚMERO DA ND (Nota Débito)
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # Procura por "ND" seguido de números
                # Exemplo: "ND: 0917167397" ou "ND 0917167397"
                nd_match = re.search(
                    r'ND\s*[:\-]?\s*(\d+)',     # Padrão regex
                    texto_completo,
                    re.IGNORECASE               # Case insensitive
                )
                if nd_match:
                    dados['numero_nd'] = nd_match.group(1)
                
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # 🔍 EXTRAÇÃO 2: DATA DE VENCIMENTO
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # Procura formato DD/MM/AAAA ou DD-MM-AAAA
                data_match = re.search(
                    r'(\d{2}[/-]\d{2}[/-]\d{4})',
                    texto_completo
                )
                if data_match:
                    dados['data_vencimento'] = data_match.group(1)
                
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # 🔍 EXTRAÇÃO 3: VALORES MONETÁRIOS
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # Procura por vários formatos de valores:
                # - R$ 1.234,56
                # - € 1.234,56
                # - 1.234,56
                # - 1234.56
                
                padroes_valores = [
                    r'R\$\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)',  # R$ com formatação BR
                    r'€\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)',     # € com formatação BR
                    r'(\d{1,3}(?:\.\d{3})*,\d{2})',              # Apenas números BR
                ]
                
                for padrao in padroes_valores:
                    matches = re.findall(padrao, texto_completo)
                    
                    for match in matches:
                        # Converte string para float
                        # "1.234,56" → 1234.56
                        valor_str = match.replace('.', '').replace(',', '.')
                        
                        try:
                            valor = float(valor_str)
                            if valor > 0:  # Apenas valores positivos
                                dados['valores_encontrados'].append(valor)
                        except:
                            pass
                
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # 🔍 EXTRAÇÃO 4: QUANTIDADE (QT)
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # Procura por "QT" ou "Quantidade" seguido de número
                qt_match = re.search(
                    r'(?:QT|Quantidade)[:\s]*(\d+(?:[.,]\d+)?)',
                    texto_completo,
                    re.IGNORECASE
                )
                if qt_match:
                    dados['quantidade'] = float(
                        qt_match.group(1).replace(',', '.')
                    )
                
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                # 🧮 CÁLCULOS FINAIS
                # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                
                # Pega o maior valor encontrado como valor total
                if dados['valores_encontrados']:
                    dados['valor_total'] = max(dados['valores_encontrados'])
                    
                    # Calcula valor unitário se tiver quantidade
                    if dados['quantidade'] > 0:
                        dados['valor_unitario'] = (
                            dados['valor_total'] / dados['quantidade']
                        )
        
        except Exception as e:
            # Se der erro, loga mas não para o programa
            self.log(f"⚠️ Erro ao processar {caminho_pdf}: {e}")
        
        return dados
    
    
    def _identificar_tipo(self, caminho):
        """
        🏷️ IDENTIFICA O TIPO DE ENCARGO PELA PASTA
        
        Verifica se o caminho contém "EAT", "PENALIDADE" ou "TOP".
        
        Args:
            caminho (str): Caminho completo do arquivo
            
        Returns:
            str: "EAT", "Penalidades", "TOP" ou "Outros"
        """
        caminho_upper = caminho.upper()  # Converte para maiúsculas
        
        if 'EAT' in caminho_upper:
            return 'EAT'
        elif 'PENALIDADE' in caminho_upper:
            return 'Penalidades'
        elif 'TOP' in caminho_upper:
            return 'TOP'
        else:
            return 'Outros'
    
    
    def _extrair_empresa(self, caminho):
        """
        🏢 EXTRAI NOME DA EMPRESA DO ARQUIVO
        
        Procura por nomes conhecidos de empresas no nome do arquivo.
        
        Args:
            caminho (str): Caminho do arquivo
            
        Returns:
            str: Nome da empresa ou "N/A"
        """
        nome = os.path.basename(caminho).upper()
        
        # Lista de empresas conhecidas
        empresas_conhecidas = [
            'COPERGAS', 'AMBEV', 'CBA', 'CERVEJARIA', 'DEXCO', 'GERDAU',
            'INDORAMA', 'INGREDION', 'KLABIN', 'MONDELEZ', 'NISSIN', 'VETRUS',
            'M DIAS BRANCO', 'PETROBRAS', 'GALP'
        ]
        
        # Verifica se alguma empresa está no nome
        for empresa in empresas_conhecidas:
            if empresa in nome:
                return empresa
        
        return 'N/A'
    
    
    def _extrair_tipo_nota(self, caminho):
        """
        💳 IDENTIFICA SE É NOTA DÉBITO OU CRÉDITO
        
        Procura por "ND", "NC", "DEBITO" ou "CREDITO" no nome.
        
        Args:
            caminho (str): Caminho do arquivo
            
        Returns:
            str: "Débito", "Crédito" ou "N/A"
        """
        nome = os.path.basename(caminho).upper()
        
        if 'ND' in nome or 'DEBITO' in nome or 'DÉBITO' in nome:
            return 'Débito'
        elif 'NC' in nome or 'CREDITO' in nome or 'CRÉDITO' in nome:
            return 'Crédito'
        
        return 'N/A'
    
    
    # ═══════════════════════════════════════════════════════════════════════
    # ⚙️ FUNÇÃO PRINCIPAL DE PROCESSAMENTO
    # ═══════════════════════════════════════════════════════════════════════
    
    def processar(self):
        """
        🚀 PROCESSA TODOS OS PDFs DA PASTA SELECIONADA
        
        Esta é a função principal que:
        1. Verifica se pasta foi selecionada
        2. Percorre recursivamente todas as subpastas
        3. Processa cada PDF encontrado
        4. Atualiza a interface com os resultados
        
        💡 FLUXO:
            selecionar_pasta() → processar() → extrair_dados_pdf()
            → _mostrar_resultados()
        """
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 🔍 VALIDAÇÕES INICIAIS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        if not self.pasta_selecionada:
            messagebox.showwarning(
                "Aviso",
                "Selecione uma pasta primeiro!"
            )
            return
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📋 INICIALIZAÇÃO
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        self.log("="*60)
        self.log("🚀 INICIANDO PROCESSAMENTO")
        self.log("="*60)
        
        self.dados_processados = []         # Limpa dados anteriores
        arquivos_processados = 0
        
        # Verifica quais tipos estão marcados
        tipos_ativos = [
            tipo
            for tipo, checkbox in self.tipos_encargo.items()
            if checkbox.get()  # .get() retorna True se marcado
        ]
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 🔄 LOOP PRINCIPAL: PERCORRE TODOS OS ARQUIVOS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        # os.walk() percorre RECURSIVAMENTE todas as pastas e subpastas
        for raiz, _, ficheiros in os.walk(self.pasta_selecionada):
            
            for ficheiro in ficheiros:
                
                # Verifica se é PDF
                if ficheiro.lower().endswith('.pdf'):
                    
                    caminho_completo = os.path.join(raiz, ficheiro)
                    
                    # Identifica o tipo
                    tipo = self._identificar_tipo(caminho_completo)
                    
                    # Pula se tipo não está ativo
                    if tipo not in tipos_ativos and tipo != 'Outros':
                        continue
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # 📄 PROCESSA O PDF
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    
                    self.log(f"📄 Processando: {ficheiro}")
                    
                    # Chama função que extrai dados do PDF
                    dados_pdf = self.extrair_dados_pdf(caminho_completo)
                    
                    # Adiciona à lista de dados processados
                    self.dados_processados.append(dados_pdf)
                    
                    # Log do resultado
                    if dados_pdf['valores_encontrados']:
                        qtd = len(dados_pdf['valores_encontrados'])
                        self.log(f"   ✓ {qtd} valores encontrados")
                    else:
                        self.log(f"   ⚠️ Sem valores")
                    
                    arquivos_processados += 1
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📊 EXIBIR RESULTADOS NA INTERFACE
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        self._mostrar_resultados(arquivos_processados)
    
    
    def _mostrar_resultados(self, total_arquivos):
        """
        📊 EXIBE RESULTADOS DO PROCESSAMENTO NA INTERFACE
        
        Atualiza todas as abas com estatísticas e dados.
        
        Args:
            total_arquivos (int): Quantidade de PDFs processados
        """
        
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Nenhum PDF encontrado!")
            return
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 🧮 CALCULAR ESTATÍSTICAS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        total_geral = sum(d['valor_total'] for d in self.dados_processados)
        com_valores = len([
            d for d in self.dados_processados if d['valor_total'] > 0
        ])
        
        # Atualiza label do total
        self.lbl_total.configure(text=f"EUR {total_geral:,.2f}")
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📊 RESUMO POR TIPO
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        resumo_tipos = {}
        for d in self.dados_processados:
            tipo = d['tipo_encargo']
            
            if tipo not in resumo_tipos:
                resumo_tipos[tipo] = {'count': 0, 'total': 0}
            
            resumo_tipos[tipo]['count'] += 1
            resumo_tipos[tipo]['total'] += d['valor_total']
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📝 ATUALIZAR ABA RESUMO
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        # Limpa widgets antigos
        for widget in self.frame_resumo.winfo_children():
            widget.destroy()
        
        # Monta texto do resumo
        stats_text = f"""
📊 ESTATÍSTICAS DO PROCESSAMENTO

Total de PDFs: {total_arquivos}
PDFs com valores: {com_valores}
Valor Total: EUR {total_geral:,.2f}

📋 RESUMO POR TIPO:
"""
        
        for tipo, stats in resumo_tipos.items():
            stats_text += f"\n{tipo}:\n"
            stats_text += f"  - Arquivos: {stats['count']}\n"
            stats_text += f"  - Total: EUR {stats['total']:,.2f}\n"
        
        # Exibe label com resumo
        ctk.CTkLabel(
            self.frame_resumo,
            text=stats_text,
            font=("Consolas", 13),
            justify="left"
        ).pack(pady=20, padx=20, anchor="w")
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📋 ATUALIZAR ABA DADOS DETALHADOS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        self._mostrar_dados_detalhados()
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # ✅ LOG FINAL
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        self.log("="*60)
        self.log(f"✅ CONCLUÍDO - {total_arquivos} arquivos")
        self.log("="*60)
        
        messagebox.showinfo(
            "Sucesso",
            f"Processados {total_arquivos} PDFs!\nTotal: EUR {total_geral:,.2f}"
        )
    
    
    def _mostrar_dados_detalhados(self):
        """
        📋 EXIBE TABELA COM DADOS DETALHADOS
        
        Cria uma tabela visual com todos os registros processados.
        Limita a 50 registros para não sobrecarregar a interface.
        """
        
        # Limpa widgets antigos
        for widget in self.frame_dados.winfo_children():
            widget.destroy()
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📋 CABEÇALHO DA TABELA
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        header = ctk.CTkFrame(self.frame_dados, fg_color="#2c3e50")
        header.pack(fill="x", pady=(0, 5))
        
        # Define colunas e larguras
        colunas = [
            ("Tipo", 80),
            ("Empresa", 150),
            ("Nota", 80),
            ("Nº", 100),
            ("Vencimento", 100),
            ("Valor Total", 120),
            ("QT", 80),
            ("Valor Unit.", 100)
        ]
        
        # Cria labels do cabeçalho
        for texto, largura in colunas:
            ctk.CTkLabel(
                header,
                text=texto,
                width=largura,
                font=("Roboto", 11, "bold")
            ).pack(side="left", padx=2)
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 📊 LINHAS DE DADOS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        
        # Limita a 50 para performance
        for d in self.dados_processados[:50]:
            
            row = ctk.CTkFrame(self.frame_dados, fg_color="#34495e")
            row.pack(fill="x", pady=1)
            
            # Valores a serem exibidos
            valores = [
                (d['tipo_encargo'], 80),
                (d['empresa'], 150),
                (d['nota_tipo'], 80),
                (d['numero_nd'], 100),
                (d['data_vencimento'], 100),
                (f"{d['valor_total']:.2f}", 120),
                (f"{d['quantidade']:.2f}", 80),
                (f"{d['valor_unitario']:.2f}", 100)
            ]
            
            # Cria labels para cada valor
            for valor, largura in valores:
                ctk.CTkLabel(
                    row,
                    text=str(valor),
                    width=largura,
                    font=("Roboto", 10)
                ).pack(side="left", padx=2)
    
    
    # ═══════════════════════════════════════════════════════════════════════
    # 💾 FUNÇÃO DE SALVAMENTO NO BANCO DE DADOS
    # ═══════════════════════════════════════════════════════════════════════
    
    def salvar_db(self):
        """
        💾 SALVA DADOS NO BANCO SQLite
        
        Cria banco de dados SQLite e insere todos os registros.
        O banco é salvo na mesma pasta dos PDFs.
        
        💡 ESTRUTURA DO BANCO:
            Tabela: dados_ret
            Campos: id, tipo_encargo, empresa, nota_tipo, numero_nd,
                    data_vencimento, valor_total, quantidade, 
                    valor_unitario, arquivo, caminho, data_processamento
        """
        
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        try:
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 🔧 CRIAR/CONECTAR BANCO DE DADOS
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            db_path = os.path.join(self.pasta_selecionada, 'RET_dados.db')
            conexao = sqlite3.connect(db_path)
            cursor = conexao.cursor()
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📋 CRIAR TABELA (SE NÃO EXISTIR)
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS dados_ret (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo_encargo TEXT,
                    empresa TEXT,
                    nota_tipo TEXT,
                    numero_nd TEXT,
                    data_vencimento TEXT,
                    valor_total REAL,
                    quantidade REAL,
                    valor_unitario REAL,
                    arquivo TEXT,
                    caminho TEXT,
                    data_processamento TEXT
                )
            ''')
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 💾 INSERIR DADOS
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for d in self.dados_processados:
                cursor.execute('''
                    INSERT INTO dados_ret (
                        tipo_encargo, empresa, nota_tipo, numero_nd,
                        data_vencimento, valor_total, quantidade, valor_unitario,
                        arquivo, caminho, data_processamento
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    d['tipo_encargo'],
                    d['empresa'],
                    d['nota_tipo'],
                    d['numero_nd'],
                    d['data_vencimento'],
                    d['valor_total'],
                    d['quantidade'],
                    d['valor_unitario'],
                    d['arquivo'],
                    d['caminho'],
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            
            # Confirma as mudanças
            conexao.commit()
            conexao.close()
            
            self.log(f"✓ Dados salvos: {db_path}")
            messagebox.showinfo("Sucesso", f"Dados salvos!\n{db_path}")
            
        except Exception as e:
            self.log(f"❌ Erro ao salvar: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar: {e}")
    
    
    # ═══════════════════════════════════════════════════════════════════════
    # 📊 FUNÇÃO DE EXPORTAÇÃO PARA EXCEL
    # ═══════════════════════════════════════════════════════════════════════
    
    def exportar_excel(self):
        """
        📊 EXPORTA DADOS PARA EXCEL FORMATADO
        
        Cria arquivo Excel profissional com 3 abas:
        1. Dados Completos
        2. Resumo por Tipo
        3. Resumo Geral
        
        Usa OpenPyXL para formatação avançada (cores, bordas, etc.)
        
        💡 FORMATAÇÃO:
            - Cabeçalhos: Fundo azul + texto branco + negrito
            - Bordas: Em todas as células
            - Números: Formato #,##0.00
            - Larguras: Ajustadas automaticamente
        """
        
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        try:
            excel_path = os.path.join(
                self.pasta_selecionada,
                'RET_Relatorio.xlsx'
            )
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📋 CRIAR DATAFRAME COM DADOS
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            df = pd.DataFrame([{
                'Tipo de Encargo': d['tipo_encargo'],
                'Empresa': d['empresa'],
                'Nota Debito/Credito': d['nota_tipo'],
                'Nº': d['numero_nd'],
                'Data Vencimento': d['data_vencimento'],
                'Valor Total': d['valor_total'],
                'QT': d['quantidade'],
                'Valor Unitario': d['valor_unitario'],
                'Arquivo': d['arquivo']
            } for d in self.dados_processados])
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 🎨 CRIAR WORKBOOK E ESTILOS
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            wb = Workbook()
            ws_dados = wb.active
            ws_dados.title = "Dados Completos"
            
            # Definir estilos
            header_fill = PatternFill(
                start_color="1F4788",
                end_color="1F4788",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📊 ABA 1: DADOS COMPLETOS
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    
                    if r_idx == 1:  # Cabeçalho
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        # Formatar colunas numéricas
                        if c_idx in [6, 7, 8]:  # Valor, QT, Unit
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
            
            # Ajustar larguras das colunas
            ws_dados.column_dimensions['A'].width = 20
            ws_dados.column_dimensions['B'].width = 25
            ws_dados.column_dimensions['C'].width = 20
            ws_dados.column_dimensions['D'].width = 15
            ws_dados.column_dimensions['E'].width = 18
            ws_dados.column_dimensions['F'].width = 15
            ws_dados.column_dimensions['G'].width = 12
            ws_dados.column_dimensions['H'].width = 15
            ws_dados.column_dimensions['I'].width = 40
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📊 ABA 2: RESUMO POR TIPO
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            ws_resumo = wb.create_sheet("Resumo por Tipo")
            
            resumo = df.groupby('Tipo de Encargo').agg({
                'Valor Total': 'sum',
                'QT': 'sum',
                'Arquivo': 'count'
            }).rename(columns={
                'Arquivo': 'Quantidade de Arquivos'
            }).reset_index()
            
            for r_idx, row in enumerate(
                dataframe_to_rows(resumo, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if c_idx > 1 and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 18
            ws_resumo.column_dimensions['C'].width = 15
            ws_resumo.column_dimensions['D'].width = 25
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📊 ABA 3: RESUMO GERAL
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            ws_geral = wb.create_sheet("Resumo Geral")
            
            total_geral = df['Valor Total'].sum()
            total_qt = df['QT'].sum()
            total_arquivos = len(df)
            
            dados_geral = [
                ['RESUMO GERAL DO PROCESSAMENTO', ''],
                ['', ''],
                ['Metrica', 'Valor'],
                ['Total de PDFs Processados', total_arquivos],
                ['Quantidade Total (QT)', total_qt],
                ['Valor Total (EUR)', total_geral],
                ['', ''],
                ['Data do Processamento', 
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            for r_idx, row in enumerate(dados_geral, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_geral.cell(row=r_idx, column=c_idx, value=value)
                    
                    if r_idx == 1:
                        cell.font = Font(bold=True, size=16, color="1F4788")
                        ws_geral.merge_cells('A1:B1')
                    elif r_idx == 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        cell.alignment = Alignment(
                            horizontal='left',
                            vertical='center'
                        )
                        if c_idx == 2 and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            ws_geral.column_dimensions['A'].width = 30
            ws_geral.column_dimensions['B'].width = 25
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 💾 SALVAR ARQUIVO
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            wb.save(excel_path)
            
            self.log(f"✓ Excel criado: {excel_path}")
            messagebox.showinfo("Sucesso", f"Excel exportado!\n{excel_path}")
            
        except Exception as e:
            self.log(f"❌ Erro ao exportar: {e}")
            messagebox.showerror("Erro", f"Erro ao exportar: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# 🚀 EXECUÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    """
    🎬 PONTO DE ENTRADA DO PROGRAMA
    
    Este bloco só executa quando o arquivo é rodado diretamente
    (não quando é importado como módulo).
    
    Cria uma instância da classe SistemaRET e inicia o loop principal.
    """
    
    # Criar janela principal
    app = SistemaRET()
    
    # Iniciar loop de eventos (mantém janela aberta)
    # 💡 Tudo acontece dentro deste loop: cliques, digitação, etc.
    app.mainloop()


"""
════════════════════════════════════════════════════════════════════════════════
    📚 CONCEITOS IMPORTANTES PARA ESTUDO
════════════════════════════════════════════════════════════════════════════════

1️⃣ POO (Programação Orientada a Objetos):
   - Classe: Molde para criar objetos (SistemaRET)
   - Herança: SistemaRET herda de ctk.CTk
   - self: Referência ao próprio objeto
   - __init__: Construtor (método especial)

2️⃣ GUI (Interface Gráfica):
   - Widgets: Componentes visuais (Button, Label, Frame)
   - Pack: Sistema de layout (organiza widgets)
   - Command: Função chamada ao clicar botão
   - Bind: Associa evento a função

3️⃣ Processamento de PDFs:
   - PDFPlumber: Extrai texto de PDFs
   - Regex: Encontra padrões específicos
   - Parsing: Transformar texto em dados estruturados

4️⃣ Manipulação de Dados:
   - Pandas DataFrame: Tabela de dados
   - Lista de dicionários: Estrutura flexível
   - Agregação: Agrupar e somar dados

5️⃣ Banco de Dados:
   - SQLite: Banco local leve
   - SQL: Linguagem de consulta
   - CRUD: Create, Read, Update, Delete

6️⃣ Arquivos e Caminhos:
   - os.path: Manipular caminhos
   - os.walk: Percorrer pastas recursivamente
   - filedialog: Abrir diálogos nativos

7️⃣ Excel Avançado:
   - OpenPyXL: Criar/formatar Excel
   - Estilos: Fontes, cores, bordas
   - Múltiplas abas: Organização

════════════════════════════════════════════════════════════════════════════════
    🎯 DESAFIOS PARA PRATICAR
════════════════════════════════════════════════════════════════════════════════

1. Adicione um botão para limpar todos os dados
2. Implemente filtro por data de vencimento
3. Adicione gráficos no Excel (charts)
4. Crie função de busca por empresa
5. Adicione exportação para CSV
6. Implemente modo claro/escuro alternável
7. Adicione progresso bar no processamento
8. Crie histórico de processamentos
9. Adicione validação de dados
10. Implemente edição inline de dados

════════════════════════════════════════════════════════════════════════════════
"""
